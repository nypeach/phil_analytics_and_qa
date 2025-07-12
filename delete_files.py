"""
Delete Files Script

This script removes rows from Excel files and objects from JSON files based on
the files_to_delete list. It maintains Excel TEXT formatting by treating all
columns as strings using openpyxl.

Set the filename in the main block at the bottom and run directly in your IDE.
"""

import os
import sys
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import NamedStyle
from files_to_delete import files_to_delete


class FileDeleter:
    """
    Handles deletion of rows from Excel files and objects from JSON files
    while maintaining proper formatting.
    """

    def __init__(self, folder_name: str):
        """
        Initialize the file deleter.

        Args:
            folder_name (str): Name of the folder to process
        """
        self.folder_name = folder_name
        self.data_path = Path("data")
        self.input_path = self.data_path / "input" / folder_name
        self.files_to_delete = set(files_to_delete)  # Convert to set for faster lookup
        self.json_to_delete = self._create_json_delete_list()

        # Statistics
        self.excel_files_processed = 0
        self.json_files_processed = 0
        self.excel_rows_deleted = 0
        self.json_objects_deleted = 0

    def _create_json_delete_list(self) -> set:
        """
        Create the JSON delete list by transforming the files_to_delete list.
        Keeps the first, third, and fourth parts (WS_ID, AMT, CHK_NBR).

        Returns:
            set: Set of transformed identifiers for JSON deletion
        """
        json_to_delete = set()

        for file_identifier in self.files_to_delete:
            parts = file_identifier.split('_')

            if len(parts) >= 6:  # Ensure we have all required parts
                # Keep parts: WS_ID (0), AMT (2), CHK_NBR (3)
                json_identifier = f"{parts[0]}_{parts[2]}_{parts[3]}"
                json_to_delete.add(json_identifier)
                print(f"   JSON transform: {file_identifier} -> {json_identifier}")
            else:
                print(f"   ⚠️ Warning: Malformed file identifier: {file_identifier}")

        return json_to_delete

    def process_files(self) -> None:
        """
        Main method to process all Excel and JSON files in the specified folder.
        """
        print(f"🗑️ Starting file deletion process for folder: {self.folder_name}")
        print(f"📁 Processing folder: {self.input_path}")
        print(f"🎯 Files to delete: {len(self.files_to_delete)}")
        print(f"🎯 JSON objects to delete: {len(self.json_to_delete)}")

        if not self.input_path.exists():
            print(f"❌ Error: Input folder does not exist: {self.input_path}")
            return

        if not self.files_to_delete:
            print("ℹ️ No files to delete specified. Exiting.")
            return

        # Process Excel files
        self._process_excel_files()

        # Process JSON files
        self._process_json_files()

        # Print summary
        self._print_summary()

    def _process_excel_files(self) -> None:
        """Process all Excel files in the input folder."""
        print(f"\n📊 Processing Excel files...")

        excel_files = list(self.input_path.glob("*.xlsx"))
        excel_files = [f for f in excel_files if not f.name.startswith("~$")]  # Skip temp files

        if not excel_files:
            print("ℹ️ No Excel files found in the folder.")
            return

        print(f"📋 Found {len(excel_files)} Excel files to process")

        for excel_file in excel_files:
            print(f"   📄 Processing: {excel_file.name}")
            self._process_single_excel_file(excel_file)
            self.excel_files_processed += 1

    def _process_single_excel_file(self, file_path: Path) -> None:
        """
        Process a single Excel file, removing rows where the File column
        matches items in files_to_delete.

        Args:
            file_path (Path): Path to the Excel file
        """
        try:
            # Load workbook with data_only=False to preserve text formatting
            workbook = openpyxl.load_workbook(file_path, data_only=False)

            rows_deleted_in_file = 0

            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows_deleted_in_sheet = self._process_worksheet(worksheet)
                rows_deleted_in_file += rows_deleted_in_sheet

            # Save the workbook if any rows were deleted
            if rows_deleted_in_file > 0:
                # Create a backup first
                backup_path = file_path.with_suffix('.xlsx.backup')
                if not backup_path.exists():
                    import shutil
                    shutil.copy2(file_path, backup_path)
                    print(f"      💾 Backup created: {backup_path.name}")

                # Save the modified workbook
                workbook.save(file_path)
                print(f"      ✅ Deleted {rows_deleted_in_file} rows from {file_path.name}")
                self.excel_rows_deleted += rows_deleted_in_file
            else:
                print(f"      ℹ️ No rows to delete in {file_path.name}")

            workbook.close()

        except Exception as e:
            print(f"      ❌ Error processing {file_path.name}: {e}")

    def _process_worksheet(self, worksheet) -> int:
        """
        Process a single worksheet, removing rows where the File column
        matches items in files_to_delete.

        Args:
            worksheet: openpyxl worksheet object

        Returns:
            int: Number of rows deleted
        """
        if worksheet.max_row <= 1:  # No data rows
            return 0

        # Find the File column index
        file_column_index = None
        header_row = 1

        for col_idx, cell in enumerate(worksheet[header_row], start=1):
            if cell.value and str(cell.value).strip().lower() == "file":
                file_column_index = col_idx
                break

        if file_column_index is None:
            print(f"      ⚠️ Warning: 'File' column not found in sheet {worksheet.title}")
            return 0

        # Collect rows to delete (working backwards to avoid index issues)
        rows_to_delete = []

        for row_idx in range(2, worksheet.max_row + 1):  # Start from row 2 (skip header)
            file_cell = worksheet.cell(row=row_idx, column=file_column_index)
            file_value = str(file_cell.value).strip() if file_cell.value else ""

            if file_value in self.files_to_delete:
                rows_to_delete.append(row_idx)

        # Delete rows in reverse order to maintain correct indices
        rows_deleted = 0
        for row_idx in reversed(rows_to_delete):
            worksheet.delete_rows(row_idx)
            rows_deleted += 1

        return rows_deleted

    def _process_json_files(self) -> None:
        """Process all JSON files in the input folder."""
        print(f"\n📄 Processing JSON files...")

        json_files = list(self.input_path.glob("*.json"))

        if not json_files:
            print("ℹ️ No JSON files found in the folder.")
            return

        print(f"📋 Found {len(json_files)} JSON files to process")

        for json_file in json_files:
            print(f"   📄 Processing: {json_file.name}")
            self._process_single_json_file(json_file)
            self.json_files_processed += 1

    def _process_single_json_file(self, file_path: Path) -> None:
        """
        Process a single JSON file, removing objects where the key
        matches items in json_to_delete.

        Args:
            file_path (Path): Path to the JSON file
        """
        try:
            # Load JSON data
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if not isinstance(data, dict):
                print(f"      ⚠️ Warning: JSON file is not a dictionary: {file_path.name}")
                return

            # Find keys to delete
            keys_to_delete = []
            for key in data.keys():
                if key in self.json_to_delete:
                    keys_to_delete.append(key)

            # Delete the objects
            objects_deleted = 0
            for key in keys_to_delete:
                del data[key]
                objects_deleted += 1
                print(f"      🗑️ Deleted object: {key}")

            # Save the modified JSON if any objects were deleted
            if objects_deleted > 0:
                # Create a backup first
                backup_path = file_path.with_suffix('.json.backup')
                if not backup_path.exists():
                    import shutil
                    shutil.copy2(file_path, backup_path)
                    print(f"      💾 Backup created: {backup_path.name}")

                # Save the modified JSON
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)

                print(f"      ✅ Deleted {objects_deleted} objects from {file_path.name}")
                self.json_objects_deleted += objects_deleted
            else:
                print(f"      ℹ️ No objects to delete in {file_path.name}")

        except json.JSONDecodeError as e:
            print(f"      ❌ JSON decode error in {file_path.name}: {e}")
        except Exception as e:
            print(f"      ❌ Error processing {file_path.name}: {e}")

    def _print_summary(self) -> None:
        """Print a summary of the deletion process."""
        print(f"\n📊 Deletion Summary:")
        print(f"   📁 Folder processed: {self.folder_name}")
        print(f"   📊 Excel files processed: {self.excel_files_processed}")
        print(f"   📄 JSON files processed: {self.json_files_processed}")
        print(f"   🗑️ Excel rows deleted: {self.excel_rows_deleted}")
        print(f"   🗑️ JSON objects deleted: {self.json_objects_deleted}")

        if self.excel_rows_deleted > 0 or self.json_objects_deleted > 0:
            print(f"   ✅ Deletion process completed successfully!")
            print(f"   💾 Backup files were created for modified files")
        else:
            print(f"   ℹ️ No items were deleted (none found matching the delete list)")


def main(folder_name: str):
    """Main function to run the file deletion script."""
    try:
        deleter = FileDeleter(folder_name)
        deleter.process_files()

    except KeyboardInterrupt:
        print(f"\n⚠️ Process interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    filename = "Regence"
    main(filename)