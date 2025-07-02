"""
PHIL Analytics and QA Library - Excel File Combiner with JSON Support

This module replicates the exact logic from combine_xlsx_files.py but returns
a DataFrame instead of saving to a file. Also includes JSON combination functionality.
"""

import os
import pandas as pd
import json
from openpyxl import load_workbook
from pathlib import Path
from typing import Optional, Dict, List
from .exceptions import FileNotFoundError, ValidationError, DataProcessingError


class ExcelCombiner:
    """
    Combines multiple Excel files from a specified folder using the exact same
    logic as the original combine_xlsx_files.py.

    This preserves all formatting and handles headers the same way as the original,
    but returns a DataFrame instead of saving to Excel.
    """

    def __init__(self, input_folder: str, max_files: int = None, save_combined: bool = True, output_folder: str = None):
        """
        Initialize the ExcelCombiner.

        Args:
            input_folder (str): Path to the folder containing Excel files to combine
            max_files (int, optional): Maximum number of files to process (for testing)
            save_combined (bool): Whether to save a _combined.xlsx file for testing
            output_folder (str, optional): Path to the output folder for saving combined file

        Raises:
            FileNotFoundError: If the input folder doesn't exist
        """
        print(f"🔧 Initializing Excel combiner for folder: {input_folder}")
        if max_files:
            print(f"   🧪 Test mode: Limited to {max_files} files")
        if save_combined:
            print(f"   💾 Will save combined file for testing")

        if not os.path.exists(input_folder):
            raise FileNotFoundError(
                input_folder,
                file_type="input folder",
                expected_location="Current working directory or specified path"
            )

        self.input_folder = input_folder
        self.max_files = max_files
        self.save_combined = save_combined
        self.output_folder = output_folder
        self.combined_data = None
        self.file_count = 0
        self.total_rows = 0

        print(f"✅ Excel combiner initialized successfully")

    def get_excel_files(self) -> list:
        """
        Get list of Excel files - exact logic from original combine_xlsx_files.py

        Returns:
            list: List of Excel file names (.xlsx files, excluding temporary files)
        """
        print(f"📁 Scanning folder for Excel files...")

        excel_files = []
        for file_name in os.listdir(self.input_folder):
            if file_name.endswith(".xlsx") and not file_name.startswith("~$"):
                excel_files.append(file_name)

        if not excel_files:
            raise ValidationError(
                f"No Excel files found in folder: {self.input_folder}",
                validation_type="file_discovery",
                expected="*.xlsx files",
                actual="no xlsx files found"
            )

        # Apply file limit for testing
        if self.max_files and len(excel_files) > self.max_files:
            excel_files = excel_files[:self.max_files]
            print(f"🧪 Test mode: Limited to first {self.max_files} files")

        print(f"📋 Found {len(excel_files)} Excel files to process")
        for i, file_name in enumerate(excel_files, 1):
            print(f"   {i}. {file_name}")

        return excel_files

    def combine_files(self) -> pd.DataFrame:
        """
        Combine Excel files using the exact same logic as combine_xlsx_files.py
        but return a DataFrame instead of saving to Excel.

        Returns:
            pd.DataFrame: Combined data with all original formatting preserved as text
        """
        print(f"🚀 Starting file combination process (replicating combine_xlsx_files.py)...")

        # Get list of Excel files
        excel_files = self.get_excel_files()

        # Initialize variables exactly like the original
        all_data = []
        first_file = True
        expected_headers = []

        print(f"🔄 Processing {len(excel_files)} files...")

        for file_name in excel_files:
            if file_name.endswith(".xlsx") and not file_name.startswith("~$"):
                file_path = os.path.join(self.input_folder, file_name)
                print(f"📄 Processing: {file_name}")

                try:
                    # Use openpyxl exactly like the original - with data_only=False
                    wb = load_workbook(file_path, data_only=False)

                    for sheet in wb.worksheets:
                        sheet_data = []

                        for i, row in enumerate(sheet.iter_rows(), start=1):
                            row_data = []

                            # Extract cell values exactly like original
                            for cell in row:
                                row_data.append(cell.value)

                            # Handle headers exactly like original
                            if first_file and i == 1:
                                expected_headers = row_data.copy()
                                sheet_data.append(row_data)
                                continue

                            # Skip header rows from subsequent files
                            if not first_file and i == 1:
                                headers = row_data.copy()
                                if headers != expected_headers:
                                    print(f"⚠️ Header mismatch in file: {file_name}")
                                continue

                            # Add the data row
                            sheet_data.append(row_data)

                        # Add this sheet's data to all_data
                        all_data.extend(sheet_data)

                    first_file = False
                    self.file_count += 1
                    print(f"   ✅ Successfully processed {file_name}")

                except Exception as e:
                    print(f"   ❌ Failed to process {file_name}: {e}")
                    continue

        if not all_data:
            raise DataProcessingError(
                "No files were successfully processed",
                operation="file_combination",
                files_attempted=len(excel_files)
            )

        # Convert to DataFrame exactly preserving the original logic
        print(f"🔗 Converting combined data to DataFrame...")

        try:
            # Create DataFrame with first row as headers, rest as data
            if len(all_data) > 0:
                headers = all_data[0]
                data_rows = all_data[1:] if len(all_data) > 1 else []

                # Create DataFrame ensuring all data is treated as strings (like Excel TEXT)
                self.combined_data = pd.DataFrame(data_rows, columns=headers)

                # Convert all columns to string to preserve Excel TEXT formatting
                for col in self.combined_data.columns:
                    self.combined_data[col] = self.combined_data[col].astype(str)

                # Replace 'None' strings with empty strings
                self.combined_data = self.combined_data.replace('None', '')
                self.combined_data = self.combined_data.fillna('')

                self.total_rows = len(self.combined_data)

                print(f"✅ File combination completed successfully!")
                print(f"   📊 Total files processed: {self.file_count}")
                print(f"   📈 Total rows combined: {self.total_rows:,}")
                print(f"   📋 Total columns: {len(self.combined_data.columns)}")

                # Show sample of the File column to verify it has the right data
                if 'File' in self.combined_data.columns:
                    sample_files = self.combined_data['File'].unique()[:3]
                    print(f"   🔍 Sample File column values: {sample_files.tolist()}")

                # Save combined file if requested
                if self.save_combined:
                    self._save_combined_file()

                return self.combined_data
            else:
                raise DataProcessingError("No data found in combined files")

        except Exception as e:
            raise DataProcessingError(
                f"Failed to create DataFrame from combined data: {e}",
                operation="dataframe_creation"
            )

    def _save_combined_file(self) -> None:
        """Save the combined data to a _combined.xlsx file in the output folder."""
        if self.combined_data is None:
            return

        # Extract payer name from input folder
        payer_name = os.path.basename(self.input_folder.rstrip('/\\'))

        # Determine output directory
        if self.output_folder:
            # Use the provided output folder
            output_dir = self.output_folder
            # Create output folder if it doesn't exist
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
        else:
            # Fallback to parent directory of input folder
            output_dir = os.path.dirname(self.input_folder)

        combined_filename = f"{payer_name}_combined.xlsx"
        combined_file_path = os.path.join(output_dir, combined_filename)

        print(f"💾 Saving combined file: {combined_filename}")
        print(f"   📁 Output directory: {output_dir}")

        try:
            # Use openpyxl engine to maintain text formatting
            with pd.ExcelWriter(combined_file_path, engine='openpyxl') as writer:
                self.combined_data.to_excel(writer, sheet_name='Sheet1', index=False)

                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                # Format all columns as text to preserve formatting
                from openpyxl.utils import get_column_letter

                for col_idx, col_name in enumerate(self.combined_data.columns, 1):
                    col_letter = get_column_letter(col_idx)

                    # Set column format to text for all cells in this column
                    for row in range(1, len(self.combined_data) + 2):  # +2 for header and 1-indexing
                        cell = worksheet[f"{col_letter}{row}"]
                        cell.number_format = '@'  # Text format

                # Bold the header row
                for cell in worksheet[1]:
                    cell.font = cell.font.copy(bold=True)

                # Freeze the top row
                worksheet.freeze_panes = 'A2'

            print(f"   ✅ Combined file saved successfully!")
            print(f"   📁 Full path: {combined_file_path}")
            print(f"   📊 Rows saved: {len(self.combined_data):,}")

        except Exception as e:
            print(f"   ⚠️ Warning: Could not save combined file: {e}")

    def get_file_summary(self) -> dict:
        """
        Get summary statistics about the combined data.

        Returns:
            dict: Summary information including file count, row count, columns, etc.
        """
        if self.combined_data is None:
            return {"status": "No data combined yet"}

        # Get file distribution if File column exists
        file_counts = {}
        if 'File' in self.combined_data.columns:
            file_counts = self.combined_data['File'].value_counts().to_dict()

        summary = {
            "total_files": self.file_count,
            "total_rows": self.total_rows,
            "total_columns": len(self.combined_data.columns),
            "file_distribution": file_counts,
            "columns": list(self.combined_data.columns),
            "memory_usage_mb": round(self.combined_data.memory_usage(deep=True).sum() / 1024 / 1024, 2)
        }

        return summary

    def save_to_file(self, output_file: str) -> None:
        """
        Save the combined data to an Excel file (optional - for debugging).

        Args:
            output_file (str): Path for the output Excel file
        """
        if self.combined_data is None:
            raise DataProcessingError(
                "No data to save. Run combine_files() first.",
                operation="file_save"
            )

        print(f"💾 Saving combined data to: {output_file}")

        try:
            self.combined_data.to_excel(output_file, index=False)
            print(f"✅ Combined data saved successfully!")
            print(f"   📁 File location: {output_file}")
            print(f"   📊 Rows saved: {len(self.combined_data):,}")

        except Exception as e:
            raise DataProcessingError(
                f"Failed to save Excel file: {e}",
                operation="excel_save",
                output_file=output_file
            )


class JsonCombiner:
    """
    Combines multiple JSON files from a specified folder into a single combined JSON file.

    This class handles JSON files that contain remittance data and combines them
    into a single file for easier processing and updates to "Not Posted" services.
    """

    def __init__(self, input_folder: str, output_folder: str, file_name: str):
        """
        Initialize the JSON combiner.

        Args:
            input_folder (str): Path to folder containing JSON files
            output_folder (str): Path to folder for output
            file_name (str): Base name for combined file (without extension)
        """
        print(f"🔧 Initializing JSON combiner for folder: {input_folder}")

        self.input_folder = input_folder
        self.output_folder = output_folder
        self.file_name = file_name
        self.combined_data = {}
        self.file_count = 0

        # Ensure folders exist
        if not os.path.exists(input_folder):
            raise FileNotFoundError(
                input_folder,
                file_type="JSON input folder",
                expected_location="Current working directory or specified path"
            )

        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
            print(f"📁 Created output folder: {output_folder}")

        print(f"✅ JSON combiner initialized successfully")

    def get_json_files(self) -> List[str]:
        """
        Get list of JSON files in the input folder.

        Returns:
            List[str]: List of JSON file names
        """
        print(f"📁 Scanning folder for JSON files...")

        json_files = []
        for file_name in os.listdir(self.input_folder):
            if file_name.endswith(".json") and not file_name.startswith("."):
                json_files.append(file_name)

        if not json_files:
            print(f"⚠️ No JSON files found in folder: {self.input_folder}")
            return []

        print(f"📋 Found {len(json_files)} JSON files to process")
        for i, file_name in enumerate(json_files, 1):
            print(f"   {i}. {file_name}")

        return json_files

    def combine_json_files(self) -> Dict:
        """
        Combine all JSON files in the input folder.

        Returns:
            Dict: Combined JSON data
        """
        print(f"🚀 Starting JSON file combination process...")

        json_files = self.get_json_files()

        if not json_files:
            print(f"ℹ️ No JSON files to combine")
            return {}

        combined_data = {}

        print(f"🔄 Processing {len(json_files)} JSON files...")

        for file_name in json_files:
            file_path = os.path.join(self.input_folder, file_name)
            print(f"📄 Processing: {file_name}")

            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    file_data = json.load(f)

                # Merge the data - each JSON file may contain multiple remittance records
                if isinstance(file_data, dict):
                    combined_data.update(file_data)
                elif isinstance(file_data, list):
                    # If it's a list, convert to dict with index as key
                    for i, item in enumerate(file_data):
                        combined_data[f"{file_name}_{i}"] = item

                self.file_count += 1
                print(f"   ✅ Successfully processed {file_name}")

            except json.JSONDecodeError as e:
                print(f"   ❌ JSON decode error in {file_name}: {e}")
                continue
            except Exception as e:
                print(f"   ❌ Failed to process {file_name}: {e}")
                continue

        self.combined_data = combined_data

        print(f"✅ JSON combination completed successfully!")
        print(f"   📊 Total files processed: {self.file_count}")
        print(f"   📈 Total remittance records: {len(combined_data)}")

        return combined_data

    def save_combined_json(self) -> str:
        """
        Save the combined JSON data to output file.

        Returns:
            str: Path to saved file
        """
        if not self.combined_data:
            print(f"⚠️ No data to save")
            return ""

        output_file = os.path.join(self.output_folder, f"{self.file_name}_combined.json")

        print(f"💾 Saving combined JSON to: {output_file}")

        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(self.combined_data, f, indent=2, ensure_ascii=False)

            print(f"   ✅ Combined JSON saved successfully!")
            print(f"   📁 File location: {output_file}")
            print(f"   📊 Records saved: {len(self.combined_data):,}")

            return output_file

        except Exception as e:
            raise DataProcessingError(
                f"Failed to save combined JSON file: {e}",
                operation="json_save",
                output_file=output_file
            )

    def run_combination(self) -> str:
        """
        Run the complete JSON combination process.

        Returns:
            str: Path to combined JSON file
        """
        self.combine_json_files()
        return self.save_combined_json()


def transform_json_service_to_data_object(json_service: Dict, json_claim: Dict) -> Dict:
    """
    Transform JSON service structure to match our data object service format.

    Args:
        json_service (Dict): Service data from JSON
        json_claim (Dict): Claim data from JSON (for context)

    Returns:
        Dict: Service object in our data format
    """
    # Build codes string from JSON adjustments
    codes_parts = []

    for adj in json_service.get("adjustments", []):
        for code, amount in adj.items():
            if code.startswith(("CO", "CR", "OA", "PI", "PR")):
                # Format: "CO253 (Description) -$amount"
                # For now, we'll use placeholder description since JSON doesn't include it
                codes_parts.append(f"{code} () -${amount}")

    codes_string = "; ".join(codes_parts)

    # Handle remarks (JSON remarks are typically empty based on your example)
    remarks = []
    for remark in json_service.get("remarks", []):
        if isinstance(remark, str):
            remarks.append(remark)
        elif isinstance(remark, dict):
            remarks.extend(remark.keys())

    # Transform to our data object format
    service_obj = {
        "clm_sts": str(json_claim.get("clm_status", "")).strip(),
        "posting_sts": "Not Posted",  # We're only updating Not Posted services
        "cpt4": str(json_service.get("proc", "")).strip(),
        "txn_status": "",  # Not available in JSON
        "description": "",  # Not available in JSON
        "bill_amt": str(json_service.get("billed", "0.00")).strip(),
        "paid_amt": str(json_service.get("prov_pd", "0.00")).strip(),
        "ded_amt": "",  # May need to calculate from adjustments
        "codes": codes_string,  # Now a formatted string
        "remarks": remarks
    }

    return service_obj


def update_service_codes_from_json(current_codes: str, json_adjustments: List[Dict]) -> str:
    """
    Update service codes string with new adjustments from JSON.
    - Add codes that aren't already present
    - Override amounts of codes with the JSON amount if the values are different

    Args:
        current_codes (str): Current codes string like "CO253 (Desc) -$0.26; CO45 (Desc) -$68.94"
        json_adjustments (List[Dict]): JSON adjustments like [{"CO253": "1532.90", "CO45": "68.94"}]

    Returns:
        str: Updated codes string
    """
    # Parse existing codes and their amounts
    existing_codes = {}
    remaining_parts = []

    if current_codes:
        parts = current_codes.split(";")
        for part in parts:
            part = part.strip()
            if part:
                # Extract code name (everything before the first space)
                space_idx = part.find(" ")
                if space_idx > 0:
                    code = part[:space_idx].strip()
                    existing_codes[code] = part  # Keep full formatted string
                else:
                    # Malformed part, keep as is
                    remaining_parts.append(part)

    # Process JSON adjustments
    json_codes = {}
    for adj in json_adjustments:
        for code, amount in adj.items():
            if code.startswith(("CO", "CR", "OA", "PI", "PR")):
                json_codes[code] = str(amount).strip()

    # Build updated codes list
    updated_parts = []

    # Handle existing codes
    for code, formatted_string in existing_codes.items():
        if code in json_codes:
            # Code exists in both - check if amount is different
            json_amount = json_codes[code]

            # Extract current amount from formatted string
            # Look for -$ pattern
            dollar_idx = formatted_string.find("-$")
            if dollar_idx > 0:
                current_amount = formatted_string[dollar_idx + 2:].strip()
                # Remove any trailing semicolon or other characters
                current_amount = current_amount.split(";")[0].strip()

                if current_amount != json_amount:
                    # Override with JSON amount, preserve description
                    desc_part = formatted_string[:dollar_idx].strip()
                    updated_parts.append(f"{desc_part} -${json_amount}")
                    print(f"   📝 Updated {code} amount: {current_amount} → {json_amount}")
                else:
                    # Keep existing
                    updated_parts.append(formatted_string)
            else:
                # Malformed existing string, use JSON format
                updated_parts.append(f"{code} () -${json_amount}")

            # Remove from json_codes so we don't add it again
            del json_codes[code]
        else:
            # Code only exists in current, keep as is
            updated_parts.append(formatted_string)

    # Add any remaining parts that weren't codes
    updated_parts.extend(remaining_parts)

    # Add new codes from JSON
    for code, amount in json_codes.items():
        updated_parts.append(f"{code} () -${amount}")
        print(f"   📝 Added new code: {code} -${amount}")

    return "; ".join(updated_parts)


def compare_and_update_service(current_service: Dict, json_service_data: Dict) -> Dict:
    """
    Compare current service with JSON data and update if different.
    Only updates services with posting_sts = "Not Posted".

    Args:
        current_service (Dict): Current service from our data object
        json_service_data (Dict): JSON service data with service, claim, remit

    Returns:
        Dict: Updated service object
    """
    # Only update if posting status is "Not Posted"
    if current_service.get("posting_sts", "").strip() != "Not Posted":
        return current_service

    json_service = json_service_data["service"]
    json_claim = json_service_data["claim"]

    # Create updated service
    updated_service = current_service.copy()

    # Update amounts if different
    json_bill_amt = str(json_service.get("billed", "0.00")).strip()
    json_paid_amt = str(json_service.get("prov_pd", "0.00")).strip()

    if updated_service.get("bill_amt", "").strip() != json_bill_amt:
        print(f"   📝 Updating bill_amt: {updated_service.get('bill_amt')} → {json_bill_amt}")
        updated_service["bill_amt"] = json_bill_amt

    if updated_service.get("paid_amt", "").strip() != json_paid_amt:
        print(f"   📝 Updating paid_amt: {updated_service.get('paid_amt')} → {json_paid_amt}")
        updated_service["paid_amt"] = json_paid_amt

    # Update codes if different
    current_codes = updated_service.get("codes", "")
    updated_codes = update_service_codes_from_json(current_codes, json_service.get("adjustments", []))

    if current_codes != updated_codes:
        print(f"   📝 Updating codes:")
        print(f"      Old: {current_codes}")
        print(f"      New: {updated_codes}")
        updated_service["codes"] = updated_codes

    return updated_service


def find_matching_json_data(encounter_num: str, claim_status: str, cpt4: str,
                          combined_json: Dict) -> Optional[Dict]:
    """
    Find matching JSON service data for a given encounter and CPT4.

    Args:
        encounter_num (str): Encounter number to match (corresponds to claim.number)
        claim_status (str): Claim status to match
        cpt4 (str): CPT4 code to match
        combined_json (Dict): Combined JSON data

    Returns:
        Optional[Dict]: Matching service data or None if not found
    """
    # Search through all remittance records
    for remit_key, remit_data in combined_json.items():
        claims = remit_data.get("claims", [])

        for claim in claims:
            # Match by claim number (corresponds to encounter number)
            claim_number = str(claim.get("number", "")).strip()
            claim_status_json = str(claim.get("clm_status", "")).strip()

            if claim_number == encounter_num and claim_status_json == claim_status:
                # Look for matching service
                for service in claim.get("services", []):
                    service_cpt4 = str(service.get("proc", "")).strip()
                    if service_cpt4 == cpt4:
                        return {
                            "service": service,
                            "claim": claim,
                            "remit": remit_data
                        }

    return None